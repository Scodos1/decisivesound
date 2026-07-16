"""
Exportable business reports - CSV / Excel / PDF - built from whichever
filters are active on the Analytics dashboard's Reports panel (date
range, booking status, event type, state, payment status). See
views.analytics_report, which is the only caller of this module.
"""
import csv
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from . import documents
from .constants import PAYMENT_STATUS_GROUPS, payment_status_for
from .models import Booking

REPORT_COLUMNS = [
    "Booking ID", "Name", "Phone", "Email", "Event Type", "Event Date", "State",
    "Guests", "Headsets", "Duration (hrs)", "Status", "Payment Status",
    "Amount Paid", "Estimated Total", "Outstanding", "Created At",
]


def filter_bookings(params):
    """params: a dict-like (typically request.GET) with optional keys
    date_from, date_to (event_date range), status, event_type, state,
    payment_status. Returns a filtered Booking queryset, ordered by
    event date."""
    qs = Booking.objects.order_by("event_date")

    date_from = (params.get("date_from") or "").strip()
    date_to = (params.get("date_to") or "").strip()
    status = (params.get("status") or "").strip()
    event_type = (params.get("event_type") or "").strip()
    state = (params.get("state") or "").strip()
    payment_status = (params.get("payment_status") or "").strip()

    if date_from:
        qs = qs.filter(event_date__gte=date_from)
    if date_to:
        qs = qs.filter(event_date__lte=date_to)
    if status:
        qs = qs.filter(status=status)
    if event_type:
        qs = qs.filter(event_type=event_type)
    if state:
        qs = qs.filter(state=state)
    if payment_status and payment_status in PAYMENT_STATUS_GROUPS:
        qs = qs.filter(status__in=PAYMENT_STATUS_GROUPS[payment_status])

    return qs


def describe_filters(params):
    """Human-readable summary of the active filters, for the PDF report
    header. Returns None if no filters are active (a plain "all bookings"
    report)."""
    labels = []
    date_from = (params.get("date_from") or "").strip()
    date_to = (params.get("date_to") or "").strip()
    if date_from or date_to:
        labels.append(f"Event date: {date_from or 'any'} to {date_to or 'any'}")
    for key, title in [
        ("status", "Status"), ("event_type", "Event type"),
        ("state", "State"), ("payment_status", "Payment status"),
    ]:
        value = (params.get(key) or "").strip()
        if value:
            labels.append(f"{title}: {value}")
    return " · ".join(labels) if labels else "All bookings"


def _report_row(booking):
    total = booking.estimated_total() if booking.full_price_known else None
    outstanding = (total - booking.amount_paid) if total is not None else None
    return [
        booking.id, booking.name, booking.phone, booking.email, booking.event_type,
        booking.event_date, booking.state, booking.guests, booking.headsets, booking.duration,
        booking.status, payment_status_for(booking), booking.amount_paid,
        total if total is not None else "TBC",
        outstanding if outstanding is not None else "TBC",
        booking.created_at.strftime("%Y-%m-%d %H:%M"),
    ]


def export_csv(bookings):
    response = HttpResponse(content_type="text/csv")
    filename = f"business_report_{timezone.localdate().isoformat()}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(REPORT_COLUMNS)
    for b in bookings:
        writer.writerow(_report_row(b))
    return response


def export_excel(bookings):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Bookings Report"
    ws.append(REPORT_COLUMNS)
    header_fill = PatternFill(start_color="FF6C4DFF", end_color="FF6C4DFF", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = header_fill

    for b in bookings:
        ws.append(_report_row(b))

    for column_cells in ws.columns:
        length = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 40)

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"business_report_{timezone.localdate().isoformat()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_pdf(bookings, filters_summary):
    bookings = list(bookings)
    rows = [
        {
            "booking": b,
            "payment_status": payment_status_for(b),
            "total": b.estimated_total() if b.full_price_known else None,
        }
        for b in bookings
    ]
    context = {
        "rows": rows,
        "filters_summary": filters_summary,
        "generated_at": timezone.now(),
        "total_bookings": len(bookings),
        "total_amount_paid": sum(b.amount_paid for b in bookings),
        "company_name": documents.COMPANY_NAME,
        "company_tagline": documents.COMPANY_TAGLINE,
    }
    pdf_bytes = documents.render_pdf("booking/documents/report_pdf.html", context)
    response = HttpResponse(pdf_bytes or b"", content_type="application/pdf")
    filename = f"business_report_{timezone.localdate().isoformat()}.pdf"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
